#!/usr/bin/env python3
"""
Workday 職缺爬蟲 (支援自動辨識台灣地點)

原理:
    幾乎所有公司的 Workday 招募網站 (*.myworkdayjobs.com) 前端都是 React,
    真正的資料來自同一組內部 JSON API:

        POST https://{tenant}.wd{N}.myworkdayjobs.com/wday/cxs/{tenant}/{site}/jobs

    所以不需要 Selenium, 直接打 API 即可。

台灣職缺怎麼判斷 (兩層策略):
    第 1 層 [優先] 讀取 API 回傳的 facets (官方的地點選單), 自動找出
        對應台灣的選項, 再用 appliedFacets 做「伺服器端篩選」。
        最準, 而且不會漏掉跨地點的職缺。
        注意: facet 的參數名稱與 id 每家公司都不一樣 (NVIDIA 是
        locationHierarchy1, Salesforce 是自訂名稱), 所以一律動態探測。

    第 2 層 [後備] 探測不到 facet 時, 改用關鍵字比對 locationsText,
        涵蓋 Taiwan / Taipei / Hsinchu / 新竹 / 台北 ... 等寫法。
        清單 API 對跨地點職缺只會給 "2 Locations", 這種會自動再打
        詳細頁, 從 location + additionalLocations 取得真實地點。

用法:
    # 先看這家公司有哪些地點可選 (含台灣自動標記)
    python3 workday_jobs.py --url <網址> --list-locations

    # 只抓台灣的職缺
    python3 workday_jobs.py --url <網址> --taiwan --csv tw_jobs.csv

    # 台灣 + 關鍵字 + 完整職務說明
    python3 workday_jobs.py --url <網址> --taiwan --search "firmware" --detail --csv out.csv
"""

import argparse
import csv
import json
import os
import re
import sys
import time
from datetime import datetime
from urllib.parse import urlparse

import requests

# 腳本自己所在的資料夾。沒指定輸出路徑時就存這裡,
# 這樣不管從哪個目錄執行 (或按 IDE 的 Run 按鈕) 檔案都會在腳本旁邊。
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

# ===========================================================================
#  要追蹤的公司清單
#  以後只要在這裡增減網址就好, 不用改指令, 直接執行就會全部跑一遍。
#  網址就到該公司職缺頁複製整段貼上, 後面的 ?xxx=yyy 查詢參數不用管
#  (程式會忽略, 改用自己偵測到的台灣篩選條件)。
# ===========================================================================
COMPANIES = [
    "https://nvidia.wd5.myworkdayjobs.com/en-US/NVIDIAExternalCareerSite",
    "https://intel.wd1.myworkdayjobs.com/External",
    "https://marvell.wd1.myworkdayjobs.com/MarvellCareers",
    "https://cadence.wd1.myworkdayjobs.com/en-US/External_Careers",
    "https://hp.wd5.myworkdayjobs.com/zh-TW/ExternalCareerSite",
    "https://hpe.wd5.myworkdayjobs.com/Jobsathpe",
    "https://analogdevices.wd1.myworkdayjobs.com/External",
]

# 直接執行 (不帶任何參數) 時套用的預設值 —— 也可以用指令列參數覆蓋
TAIWAN_ONLY = True      # 只抓台灣的職缺
SEARCH_TEXT = ""        # 送給 Workday 的關鍵字 (會連職務說明一起搜, 用來縮小抓取範圍)
TITLE_FILTER = ""       # 只保留「職稱」含這些字的職缺, 多個用 | 分隔, 例: "firmware|MCU|embedded"
FETCH_DETAIL = False    # 是否連完整職務說明一起抓 (每筆多一次請求, 會慢很多)
FETCH_JOB_TYPE = True   # 是否補上 Job Type 欄位 (Regular / Intern / New College Graduate)

PAGE_SIZE = 20  # Workday 單次查詢上限就是 20, 填更大會被忽略

HEADERS = {
    "Content-Type": "application/json",
    "Accept": "application/json",
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/124.0 Safari/537.36"
    ),
}

# ---------------------------------------------------------------------------
# 台灣地點的各種寫法。很多公司不會寫 "Taiwan", 只寫城市或園區名。
# 用 \b 詞界比對, 避免 "Tainan" 被 "Nan" 之類的片段誤中。
# ---------------------------------------------------------------------------
TAIWAN_TERMS_EN = [
    "Taiwan", "Formosa", "R.O.C.", "ROC",
    # 直轄市 / 縣市
    "Taipei", "New Taipei", "Hsinchu", "Taoyuan", "Taichung", "Tainan",
    "Kaohsiung", "Keelung", "Chiayi", "Miaoli", "Changhua", "Yunlin",
    "Nantou", "Pingtung", "Yilan", "Ilan", "Hualien", "Taitung",
    "Penghu", "Kinmen", "Matsu",
    # 科學園區 / 常見廠辦所在區
    "Zhubei", "Chupei", "Jhubei",       # 竹北
    "Zhunan", "Chunan", "Jhunan",       # 竹南
    "Hukou", "Longtan", "Lungtan",      # 湖口 / 龍潭
    "Linkou", "Neihu", "Nangang", "Xinyi", "Zhongli", "Chungli",
    "Xindian", "Banqiao", "Zhonghe", "Tucheng", "Xizhi", "Guishan",
]
TAIWAN_TERMS_ZH = [
    "台灣", "臺灣", "台北", "臺北", "新北", "新竹", "竹北", "竹南",
    "桃園", "中壢", "台中", "臺中", "台南", "臺南", "高雄", "基隆",
    "嘉義", "苗栗", "彰化", "雲林", "南投", "屏東", "宜蘭", "花蓮",
    "台東", "臺東", "內湖", "南港", "林口", "龍潭", "新店", "板橋",
    "中和", "土城", "湖口", "汐止", "龜山", "科學園區",
]

# 英文用詞界; 中文沒有詞界概念, 直接子字串比對
_TW_RE_EN = re.compile(
    r"\b(" + "|".join(re.escape(t) for t in TAIWAN_TERMS_EN) + r")\b", re.IGNORECASE
)
_TW_RE_ZH = re.compile("|".join(re.escape(t) for t in TAIWAN_TERMS_ZH))
# "TW" 只在單獨成 token 時才算 (例: "TW, Taipei"), 避免誤判
_TW_RE_CODE = re.compile(r"(?<![A-Za-z])TW(?![A-Za-z])")

# 清單 API 對跨地點職缺會回 "2 Locations" 這種字串
_MULTI_LOC_RE = re.compile(r"^\s*\d+\s+Locations?\s*$", re.IGNORECASE)

# 判斷哪些 facet 是「地點類」的
_LOCATION_FACET_RE = re.compile(r"location|country|site|region|city|geo", re.IGNORECASE)


def is_taiwan(text):
    """判斷一段地點文字是否屬於台灣。"""
    if not text:
        return False
    return bool(_TW_RE_EN.search(text) or _TW_RE_ZH.search(text) or _TW_RE_CODE.search(text))


# ---------------------------------------------------------------------------
# 網址解析
# ---------------------------------------------------------------------------
def parse_workday_url(url):
    """把使用者貼的職缺頁網址拆成 (host, tenant, site)。

    https://nvidia.wd5.myworkdayjobs.com/en-US/NVIDIAExternalCareerSite
        -> ('nvidia.wd5.myworkdayjobs.com', 'nvidia', 'NVIDIAExternalCareerSite')
    """
    parsed = urlparse(url if "://" in url else "https://" + url)
    host = parsed.netloc
    tenant = host.split(".")[0]

    parts = [p for p in parsed.path.split("/") if p]
    parts = [p for p in parts if not re.fullmatch(r"[a-z]{2}-[A-Z]{2}", p)]  # 去掉語系
    if not parts:
        raise ValueError(f"從網址找不到 site 名稱: {url}")

    return host, tenant, parts[0]


class WorkdayClient:
    def __init__(self, host, tenant, site, delay=0.5):
        self.host, self.tenant, self.site, self.delay = host, tenant, site, delay
        self.api = f"https://{host}/wday/cxs/{tenant}/{site}"
        self.session = requests.Session()
        self.session.headers.update(HEADERS)

    def query(self, offset=0, limit=PAGE_SIZE, search_text="", applied_facets=None):
        payload = {
            "appliedFacets": applied_facets or {},
            "limit": limit,
            "offset": offset,
            "searchText": search_text,
        }
        resp = self.session.post(f"{self.api}/jobs", json=payload, timeout=30)
        resp.raise_for_status()
        return resp.json()

    def detail(self, external_path):
        resp = self.session.get(f"{self.api}{external_path}", timeout=30)
        resp.raise_for_status()
        time.sleep(self.delay)
        return resp.json().get("jobPostingInfo", {})


# ---------------------------------------------------------------------------
# facet 探測
# ---------------------------------------------------------------------------
def flatten_facets(facets):
    """Workday 的 facets 可能巢狀 (例: locationMainGroup 底下包 3 組)。
    攤平成 [(facetParameter, descriptor, values), ...]。
    """
    out = []
    for f in facets or []:
        values = f.get("values") or []
        # 若 values 裡還帶 facetParameter, 代表這是一個群組, 要往下遞迴
        if values and isinstance(values[0], dict) and "facetParameter" in values[0]:
            out.extend(flatten_facets(values))
        else:
            out.append((f.get("facetParameter"), f.get("descriptor"), values))
    return out


def location_facets(facets):
    """只留下地點類的 facet。"""
    result = []
    for param, desc, values in flatten_facets(facets):
        looks_like_location = _LOCATION_FACET_RE.search(f"{param} {desc}")
        # 有些自訂 facet 名稱看不出來, 就看值本身像不像地名 (含逗號的國家/城市組合)
        if not looks_like_location and values:
            sample = " ".join(v.get("descriptor", "") for v in values[:20])
            looks_like_location = is_taiwan(sample) or "," in sample
        if looks_like_location:
            result.append((param, desc, values))
    return result


def find_taiwan_facet(facets):
    """從 facets 中找出對應台灣的篩選條件。

    回傳 (facetParameter, [id, ...], [descriptor, ...]) 或 None。
    每家公司的參數名稱與 id 都不同, 所以靠 descriptor 內容判斷。
    """
    candidates = []
    for param, desc, values in location_facets(facets):
        hits = [v for v in values if is_taiwan(v.get("descriptor", ""))]
        if not hits:
            continue
        # 國家層級的 facet (descriptor 剛好就是 "Taiwan") 最精準, 優先採用
        country_level = any(
            re.fullmatch(r"\s*(taiwan|台灣|臺灣)\s*", v["descriptor"], re.IGNORECASE)
            for v in hits
        )
        total = sum(v.get("count", 0) for v in hits)
        candidates.append((country_level, total, param, desc, hits))

    if not candidates:
        return None

    # 先挑國家層級, 再挑涵蓋職缺數最多的
    candidates.sort(key=lambda c: (c[0], c[1]), reverse=True)
    _, total, param, desc, hits = candidates[0]
    return {
        "param": param,
        "ids": [v["id"] for v in hits],
        "descriptors": [v["descriptor"] for v in hits],
        "facet_desc": desc,
        # 篩選面板上顯示的數字。注意這跟實際能取得的職缺數不一定相同,
        # Workday 的 facet 統計與查詢結果來自不同來源。
        "facet_count": total,
    }


# ---------------------------------------------------------------------------
# 抓取
# ---------------------------------------------------------------------------
def fetch_jobs(client, search_text="", max_jobs=None, applied_facets=None, quiet=False):
    """分頁抓取職缺清單。回傳 (jobs, total)。quiet=True 時不印進度。"""
    jobs, offset, total = [], 0, None

    while True:
        data = client.query(offset=offset, search_text=search_text,
                            applied_facets=applied_facets)
        if total is None:
            total = data.get("total", 0)
            if not quiet:
                print(f"符合條件共 {total} 筆", file=sys.stderr)

        postings = data.get("jobPostings", [])
        if not postings:
            break

        for p in postings:
            bullets = p.get("bulletFields") or []
            jobs.append({
                "title": p.get("title", ""),
                "job_id": bullets[0] if bullets else "",
                "location": p.get("locationsText", ""),
                "posted_on": p.get("postedOn", ""),
                "url": f"https://{client.host}/{client.site}{p.get('externalPath','')}",
                "_api_path": p.get("externalPath", ""),
            })

        offset += PAGE_SIZE
        if not quiet:
            print(f"  已抓取 {len(jobs)}/{total} ...", file=sys.stderr)

        if offset >= total or (max_jobs and len(jobs) >= max_jobs):
            break
        time.sleep(client.delay)

    return (jobs[:max_jobs] if max_jobs else jobs), (total or 0)


def resolve_real_locations(client, job):
    """清單只給 "2 Locations" 時, 打詳細頁取得真實地點清單。"""
    info = client.detail(job["_api_path"])
    locs = [info.get("location", "")] + list(info.get("additionalLocations") or [])
    locs = [x for x in locs if x]
    if locs:
        job["location"] = " | ".join(locs)
    return locs


def filter_taiwan_clientside(client, jobs):
    """後備方案: 用關鍵字比對挑出台灣職缺。

    "N Locations" 這種看不出地點的, 會額外打一次詳細頁確認。
    """
    kept = []
    for i, job in enumerate(jobs, 1):
        text = job["location"]

        if _MULTI_LOC_RE.match(text):
            print(f"  展開跨地點職缺 {i}/{len(jobs)} ...", file=sys.stderr)
            try:
                locs = resolve_real_locations(client, job)
            except Exception as e:
                print(f"    略過 (查詢失敗): {e}", file=sys.stderr)
                continue
            text = " ".join(locs)

        if is_taiwan(text):
            kept.append(job)
    return kept


def strip_html(html):
    text = re.sub(r"<[^>]+>", " ", html or "")
    return re.sub(r"\s+", " ", text).strip()


# Workday 只給相對時間 (詳細頁也一樣, 沒有精確日期)
_POSTED_N_RE = re.compile(r"(\d+)\s*\+?\s*day", re.IGNORECASE)
_UNKNOWN_DAYS = 9999          # 解析不出來的排最後


def posted_days(text):
    """把 "Posted 3 Days Ago" 這類字串換成天數, 用來排序。

    "Posted 30+ Days Ago" 是 Workday 的最舊級距 (實測 240 筆有 158 筆落在
    這一格), 一律當 30 天。同分時靠穩定排序保留 API 原本的新到舊順序。
    """
    if not text:
        return _UNKNOWN_DAYS
    low = text.lower()
    if "today" in low or "just posted" in low:
        return 0
    if "yesterday" in low:
        return 1
    m = _POSTED_N_RE.search(low)
    if m:
        return int(m.group(1))
    if "month" in low:
        m2 = re.search(r"(\d+)", low)
        return int(m2.group(1)) * 30 if m2 else 30
    return _UNKNOWN_DAYS


# 新鮮度級距。順序就是圖表由新到舊的堆疊順序。
FRESHNESS_BUCKETS = ["7天內", "8-14天", "15-30天", "30天以上"]


def freshness_bucket(days):
    if days <= 7:
        return FRESHNESS_BUCKETS[0]
    if days <= 14:
        return FRESHNESS_BUCKETS[1]
    if days < 30:
        return FRESHNESS_BUCKETS[2]
    return FRESHNESS_BUCKETS[3]


def sort_by_posted(jobs):
    """依張貼時間排序: 最新的在最上面。

    Python 的 sorted 是穩定排序, 所以「30+ 天」這一大群同分的職缺
    會維持 Workday 回傳的原始順序 —— 那個順序本身就是新到舊。
    """
    return sorted(jobs, key=lambda j: posted_days(j.get("posted_on", "")))


def compile_title_filter(pattern):
    """把 "firmware|MCU|embedded" 編成比對職稱用的正則。

    每段都會被 escape, 所以 "C++" 之類含特殊字元的關鍵字也能安全使用;
    | 是唯一的特殊符號, 代表「或」。大小寫不敏感。
    """
    if not pattern:
        return None
    parts = [p.strip() for p in pattern.split("|") if p.strip()]
    if not parts:
        return None
    return re.compile("|".join(re.escape(p) for p in parts), re.IGNORECASE)


# Workday 各家的 Job Type facet 參數名稱 (四家實測都是 workerSubType)
_JOB_TYPE_PARAMS = ("workerSubType", "workerType", "employmentType", "jobType")


def find_job_type_facet(facets):
    """找出 Job Type 這個 facet (Regular / Intern / New College Graduate ...)。"""
    flat = flatten_facets(facets)
    for param, desc, values in flat:
        if param in _JOB_TYPE_PARAMS and values:
            return param, values
    # 參數名稱不認得時, 退而求其次看顯示名稱
    for param, desc, values in flat:
        if desc and re.fullmatch(r"\s*job\s*type\s*", desc, re.IGNORECASE) and values:
            return param, values
    return None


def annotate_job_type(client, jobs, base_facets=None, search_text=""):
    """幫每筆職缺補上 Job Type 欄位。

    做法是「反查」: 對 Job Type 的每個選項各查一次, 把回傳的職缺編號標記成
    該類型。這些選項會把職缺集合切成互斥的子集, 所以總請求數約等於再抓一遍
    清單 —— 比逐筆打詳細頁 (每筆一次請求) 便宜一個數量級。
    """
    probe = client.query(limit=1, search_text=search_text, applied_facets=base_facets)
    hit = find_job_type_facet(probe.get("facets", []))
    if not hit:
        return 0
    param, values = hit

    wanted = {j["job_id"] for j in jobs}
    mapping = {}
    for v in values:
        if not wanted - mapping.keys():        # 全部標記完就不用再查了
            break
        facets = dict(base_facets or {})
        facets[param] = [v["id"]]
        found, _ = fetch_jobs(client, search_text, None, facets, quiet=True)
        for job in found:
            if job["job_id"] in wanted:
                mapping[job["job_id"]] = v["descriptor"]

    for job in jobs:
        job["job_type"] = mapping.get(job["job_id"], "")
    return len(mapping)


# Workday 未篩選查詢的 total 會被截斷在這個值
_RESULT_CAP = 2000


def report_count_gap(facet_count, total, fetched, max_jobs=None):
    """比對「篩選面板顯示的數字」與「實際取得的職缺數」, 有落差就說明原因。

    這兩個數字來自 Workday 不同的資料來源, 差 1~2 筆是正常現象:
    facet 是照統計索引算的, 查詢結果只回傳目前真正開放的職缺。
    """
    lines = []

    if facet_count is not None and facet_count != total:
        gap = facet_count - total
        if gap > 0:
            lines.append(
                f"注意: 篩選面板顯示 {facet_count} 筆, 實際可取得 {total} 筆 (差 {gap} 筆)。\n"
                f"      這不是漏抓 — facet 數字是照 Workday 統計索引算的, 會把已關閉或\n"
                f"      隱藏的職缺算進去; 查詢結果只回傳目前真正開放、可投遞的職缺。"
            )
        else:
            lines.append(
                f"注意: 篩選面板顯示 {facet_count} 筆, 實際可取得 {total} 筆 (多 {-gap} 筆)。\n"
                f"      通常是抓取期間有新職缺上架。"
            )

    if total >= _RESULT_CAP:
        lines.append(
            f"注意: Workday 單次查詢最多只回傳 {_RESULT_CAP} 筆, 目前已達上限。\n"
            f"      請用 --search 或縮小地點範圍分批抓, 否則會有遺漏。"
        )

    if max_jobs and fetched < total:
        lines.append(f"注意: 因為指定了 --max {max_jobs}, 只抓了 {fetched}/{total} 筆。")

    for line in lines:
        print(line, file=sys.stderr)
    return bool(lines)


# ---------------------------------------------------------------------------
# 輸出
# ---------------------------------------------------------------------------
# 欄位順序 + 中文表頭
COLUMNS = [
    ("company", "公司", 12),
    ("job_id", "職缺編號", 14),
    ("title", "職缺名稱", 52),
    ("location", "地點", 30),
    ("job_type", "職務類型", 24),      # Workday 官方的 "Job Type": Regular / Intern / New College Graduate ...
    ("posted_on", "張貼時間", 18),
    ("days_ago", "幾天前", 9),         # 數值, 讓 Excel 能正確重新排序 (文字排序會錯)
    ("freshness", "新鮮度", 11),       # 級距, 給圖表與樞紐分析用
    ("time_type", "全職/兼職", 12),    # 需要 --detail 才會有
    ("remote", "遠端", 12),
    ("start_date", "開始日期", 14),
    ("url", "連結", 46),
    ("description", "職務說明", 80),
]

_EXCEL_MAX_CELL = 32767            # Excel 單一儲存格字數上限
_CTRL_RE = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")   # openpyxl 不接受的控制字元


def _ordered_fields(jobs):
    """只輸出實際存在的欄位 (沒加 --detail 就不會有職務說明那幾欄)。"""
    present = set()
    for job in jobs:
        present.update(job.keys())
    return [c for c in COLUMNS if c[0] in present]


def _clean(value):
    text = "" if value is None else str(value)
    text = _CTRL_RE.sub("", text)
    return text[:_EXCEL_MAX_CELL]


def default_output_path(tenant, taiwan_only, ext):
    """預設檔名: nvidia_taiwan_20260810.xlsx, 存在腳本所在資料夾。"""
    scope = "taiwan" if taiwan_only else "all"
    name = f"{tenant}_{scope}_{datetime.now():%Y%m%d}.{ext}"
    return os.path.join(SCRIPT_DIR, name)


SUMMARY_COLUMNS = [
    ("company", "公司", 14),
    ("count", "台灣職缺數", 12),
    ("filter", "使用的篩選條件", 46),
    ("note", "備註", 24),
    ("url", "職缺網址", 60),
]


def save_xlsx(jobs, path, summary=None):
    """存成真正的 .xlsx (自動欄寬 / 凍結表頭 / 可點擊的職缺連結)。

    summary 有值時會多一張「總覽」分頁, 列出每家公司抓到幾筆。
    沒安裝 openpyxl 時回傳 False, 由呼叫端改存 CSV。
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        return False

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4472C4")

    def write_header(ws, fields):
        for col, (_, label, width) in enumerate(fields, 1):
            cell = ws.cell(row=1, column=col, value=label)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(vertical="center")
            ws.column_dimensions[get_column_letter(col)].width = width

    fields = _ordered_fields(jobs)
    wb = Workbook()

    if summary:
        sm = wb.active
        sm.title = "總覽"
        write_header(sm, SUMMARY_COLUMNS)
        for row, item in enumerate(summary, 2):
            for col, (key, _, _) in enumerate(SUMMARY_COLUMNS, 1):
                value = item.get(key)
                cell = sm.cell(row=row, column=col,
                               value=value if key == "count" else _clean(value))
                cell.alignment = Alignment(vertical="center")
        total_row = len(summary) + 2
        sm.cell(row=total_row, column=1, value="合計").font = Font(bold=True)
        sm.cell(row=total_row, column=2, value=sum(i.get("count", 0) for i in summary)).font = Font(bold=True)
        sm.freeze_panes = "A2"
        add_freshness_chart(sm, jobs, summary, first_row=total_row + 3)
        ws = wb.create_sheet("職缺")
    else:
        ws = wb.active
        ws.title = "職缺"

    write_header(ws, fields)

    for row, job in enumerate(jobs, 2):
        for col, (key, _, _) in enumerate(fields, 1):
            raw = job.get(key)
            value = raw if (key == "days_ago" and isinstance(raw, int)) else _clean(raw)
            cell = ws.cell(row=row, column=col, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=(key in ("title", "description")))
            if key == "url" and job.get("url"):
                cell.hyperlink = job["url"]
                cell.font = Font(color="0563C1", underline="single")

    ws.freeze_panes = "A2"                                  # 捲動時表頭固定
    ws.auto_filter.ref = ws.dimensions                      # 每欄加篩選箭頭
    wb.save(path)
    return True


def add_freshness_chart(sheet, jobs, summary, first_row):
    """在總覽分頁畫「各公司職缺新鮮度」堆疊長條圖。

    用 Excel 原生圖表 (不是圖片), 所以在 Excel 裡可以直接改色、換圖型,
    列印或縮放也不會糊。資料區塊寫在圖表下方, 使用者可自行調整。
    """
    from openpyxl.chart import BarChart, Reference
    from openpyxl.chart.label import DataLabelList
    from openpyxl.styles import Font

    # 依公司統計各級距筆數
    order = [i["company"] for i in summary]
    table = {co: dict.fromkeys(FRESHNESS_BUCKETS, 0) for co in order}
    for job in jobs:
        co = job.get("company")
        if co in table:
            table[co][job.get("freshness", FRESHNESS_BUCKETS[-1])] += 1

    sheet.cell(row=first_row - 1, column=1, value="圖表資料：各公司職缺新鮮度").font = Font(bold=True)
    sheet.cell(row=first_row, column=1, value="公司").font = Font(bold=True)
    for j, label in enumerate(FRESHNESS_BUCKETS, 2):
        sheet.cell(row=first_row, column=j, value=label).font = Font(bold=True)
    for i, co in enumerate(order, 1):
        sheet.cell(row=first_row + i, column=1, value=co)
        for j, label in enumerate(FRESHNESS_BUCKETS, 2):
            sheet.cell(row=first_row + i, column=j, value=table[co][label])

    last_row = first_row + len(order)
    chart = BarChart()
    chart.type = "bar"                 # 橫向, 公司名稱較長時比較好讀
    chart.grouping = "stacked"
    chart.overlap = 100                # 堆疊必須設 100, 否則 Excel 會畫成並排
    chart.title = "各公司職缺新鮮度（依張貼時間）"
    chart.y_axis.title = "職缺數"
    chart.x_axis.delete = False        # 避免某些 Excel 版本不顯示座標軸
    chart.y_axis.delete = False
    chart.height, chart.width = 9.5, 20

    data = Reference(sheet, min_col=2, max_col=1 + len(FRESHNESS_BUCKETS),
                     min_row=first_row, max_row=last_row)
    cats = Reference(sheet, min_col=1, min_row=first_row + 1, max_row=last_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)

    # 越新越亮, 30 天以上刻意用灰色淡化
    for series, color in zip(chart.series, ["02C39A", "00A896", "028090", "C2D6DA"]):
        series.graphicalProperties.solidFill = color
        series.graphicalProperties.line.solidFill = color

    chart.dataLabels = DataLabelList()
    chart.dataLabels.showVal = True

    sheet.add_chart(chart, "G2")
    return chart


def save_csv(jobs, path):
    """utf-8-sig (帶 BOM), Excel 直接雙擊開啟不會中文亂碼。"""
    fields = _ordered_fields(jobs)
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow([label for _, label, _ in fields])
        for job in jobs:
            writer.writerow([_clean(job.get(key)) for key, _, _ in fields])


def scrape_company(url, taiwan=True, search="", max_jobs=None, detail=False,
                   delay=0.5, job_type=True, title=""):
    """抓單一公司的職缺。回傳 (jobs, info)。

    info 記錄這家公司用了什麼篩選方式、抓到幾筆, 最後會彙整進總覽表。
    """
    host, tenant, site = parse_workday_url(url)
    client = WorkdayClient(host, tenant, site, delay)
    info = {"company": tenant, "url": url, "filter": "-", "count": 0, "note": ""}

    applied_facets = None
    taiwan_by_facet = False
    facet_count = None

    if taiwan:
        # 第 1 層: 探測官方的地點 facet
        probe = client.query(limit=1, search_text=search)
        hit = find_taiwan_facet(probe.get("facets", []))
        if hit:
            print(f"  使用伺服器端篩選: {hit['param']} = {hit['descriptors']}", file=sys.stderr)
            applied_facets = {hit["param"]: hit["ids"]}
            taiwan_by_facet = True
            info["filter"] = f"{hit['param']} = {', '.join(hit['descriptors'])}"
            # 只有單一 facet 值時面板數字才可直接對照;
            # 多個值 (例: 台北+新竹) 會因跨地點職缺重複計算而失真。
            if len(hit["ids"]) == 1 and not search:
                facet_count = hit["facet_count"]
        else:
            print("  找不到台灣 facet, 改用關鍵字比對模式", file=sys.stderr)
            info["filter"] = "關鍵字比對 (無 facet)"

    jobs, total = fetch_jobs(client, search, max_jobs, applied_facets)
    if report_count_gap(facet_count, total, len(jobs), max_jobs):
        info["note"] = f"面板 {facet_count} / 實際 {total}" if facet_count != total else ""

    if taiwan:
        if taiwan_by_facet:
            # facet 已經篩過了, 但清單顯示 "N Locations" 的補一下真實地點
            for job in jobs:
                if _MULTI_LOC_RE.match(job["location"]):
                    try:
                        resolve_real_locations(client, job)
                    except Exception:
                        pass
        else:
            # 第 2 層: 關鍵字比對
            before = len(jobs)
            jobs = filter_taiwan_clientside(client, jobs)
            print(f"  關鍵字篩選: {before} -> {len(jobs)} 筆", file=sys.stderr)

    # 職稱篩選放在補 Job Type / 抓詳細內容之前, 先砍掉不要的才不用白花請求
    title_re = compile_title_filter(title)
    if title_re:
        before = len(jobs)
        jobs = [j for j in jobs if title_re.search(j["title"])]
        print(f"  職稱篩選 ({title}): {before} -> {len(jobs)} 筆", file=sys.stderr)
        info["filter"] += f" + 職稱含 {title}"

    if job_type and jobs:
        n = annotate_job_type(client, jobs, applied_facets, search)
        print(f"  補上 Job Type: {n}/{len(jobs)} 筆", file=sys.stderr)

    if detail:
        for i, job in enumerate(jobs, 1):
            print(f"  抓取詳細內容 {i}/{len(jobs)} ...", file=sys.stderr)
            try:
                d = client.detail(job["_api_path"])
                job["description"] = strip_html(d.get("jobDescription"))
                job["time_type"] = d.get("timeType", "")
                job["remote"] = d.get("remoteType", "") or ""
                job["start_date"] = d.get("startDate", "")
            except Exception as e:
                print(f"    略過 (失敗): {e}", file=sys.stderr)

    for job in jobs:
        job.pop("_api_path", None)
        job["company"] = tenant
        job["days_ago"] = posted_days(job.get("posted_on", ""))
        job["freshness"] = freshness_bucket(job["days_ago"])

    jobs = sort_by_posted(jobs)
    info["count"] = len(jobs)
    return jobs, info


# ---------------------------------------------------------------------------
def cmd_list_locations(client):
    """把這個網站所有地點選項印出來, 並標記系統判定為台灣的項目。"""
    data = client.query(limit=1)
    facets = data.get("facets", [])
    locs = location_facets(facets)

    if not locs:
        print("這個網站沒有提供地點 facet, 只能用關鍵字模式 (--taiwan 會自動退回)。")
        return

    for param, desc, values in locs:
        print(f"\n=== {desc or '(無名稱)'}  [facetParameter={param}]  共 {len(values)} 項")
        for v in sorted(values, key=lambda x: -x.get("count", 0)):
            mark = "  <== 判定為台灣" if is_taiwan(v.get("descriptor", "")) else ""
            print(f"    {v.get('count', 0):>5}  {v.get('descriptor','')}{mark}")

    hit = find_taiwan_facet(facets)
    print()
    if hit:
        print(f"台灣篩選條件: {hit['param']} = {hit['descriptors']}")
        print(f"  (--taiwan 會用這組做伺服器端篩選, 面板顯示共 {hit['facet_count']} 筆)")
    else:
        print("找不到台灣的 facet -> --taiwan 會改用關鍵字比對模式")


def main():
    ap = argparse.ArgumentParser(
        description="爬取 Workday 招募網站的職缺清單。"
                    "不帶任何參數執行時, 會跑程式碼上方 COMPANIES 清單裡的所有公司。")
    ap.add_argument("--url", action="append", dest="urls",
                    help="只抓這個網址 (可重複指定多次)。不給就用 COMPANIES 清單")
    ap.add_argument("--list-locations", action="store_true",
                    help="只列出地點選項 (含台灣自動標記), 不抓職缺")
    ap.add_argument("--taiwan", dest="taiwan", action="store_true", default=None,
                    help="只保留台灣的職缺")
    ap.add_argument("--all-locations", dest="taiwan", action="store_false",
                    help="不限地點, 全部都抓")
    ap.add_argument("--search", default=None,
                    help="送給 Workday 的關鍵字, 會連職務說明一起搜 (用來縮小抓取範圍)")
    ap.add_argument("--title", default=None,
                    help='只保留職稱含這些字的職缺, 多個用 | 分隔, 例: "firmware|MCU"')
    ap.add_argument("--max", type=int, default=None, help="每家公司最多抓幾筆 (預設全抓)")
    ap.add_argument("--detail", action="store_true", default=None,
                    help="一併抓取完整職務說明 (較慢)")
    ap.add_argument("--no-job-type", dest="job_type", action="store_false", default=None,
                    help="不要補 Job Type 欄位 (可省下約一半的請求數)")
    ap.add_argument("-o", "--out",
                    help="輸出檔案路徑。不指定就自動存到腳本所在資料夾")
    ap.add_argument("--separate", action="store_true",
                    help="每家公司各存一個檔, 而不是合併成一份")
    ap.add_argument("--format", choices=["xlsx", "csv"], default="xlsx",
                    help="輸出格式 (預設 xlsx; 沒安裝 openpyxl 會自動改成 csv)")
    ap.add_argument("--json", dest="json_out", help="額外輸出一份 JSON 檔")
    ap.add_argument("--no-save", action="store_true", help="不存檔, 只印在畫面上")
    ap.add_argument("--delay", type=float, default=0.5, help="每次請求間隔秒數 (預設 0.5)")
    args = ap.parse_args()

    # 指令列沒指定的, 一律採用檔案上方的預設值
    urls = args.urls or COMPANIES
    taiwan = TAIWAN_ONLY if args.taiwan is None else args.taiwan
    search = SEARCH_TEXT if args.search is None else args.search
    detail = FETCH_DETAIL if args.detail is None else args.detail
    job_type = FETCH_JOB_TYPE if args.job_type is None else args.job_type
    title = TITLE_FILTER if args.title is None else args.title

    if not urls:
        print("COMPANIES 清單是空的, 請在程式碼上方加入至少一個網址", file=sys.stderr)
        return

    if args.list_locations:
        for url in urls:
            host, tenant, site = parse_workday_url(url)
            print(f"\n{'=' * 70}\n{tenant}  ({url})\n{'=' * 70}")
            cmd_list_locations(WorkdayClient(host, tenant, site, args.delay))
        return

    fmt = args.format
    if args.out:                        # 有給副檔名就以副檔名為準
        ext = os.path.splitext(args.out)[1].lower()
        if ext in (".xlsx", ".csv"):
            fmt = ext.lstrip(".")

    all_jobs, summary = [], []

    for i, url in enumerate(urls, 1):
        print(f"\n[{i}/{len(urls)}] {url}", file=sys.stderr)
        try:
            jobs, info = scrape_company(url, taiwan, search, args.max, detail,
                                        args.delay, job_type, title)
        except Exception as e:
            # 單一公司失敗不影響其他家, 記錄下來繼續跑
            print(f"  失敗: {e}", file=sys.stderr)
            tenant = url.split("//")[-1].split(".")[0]
            summary.append({"company": tenant, "url": url, "filter": "-",
                            "count": 0, "note": f"失敗: {e}"})
            continue

        print(f"  -> {info['count']} 筆", file=sys.stderr)
        summary.append(info)
        all_jobs.extend(jobs)

        if args.separate and jobs and not args.no_save:
            path = default_output_path(info["company"], taiwan, fmt)
            _write(jobs, path, fmt, None)

    # 螢幕上印一份總覽, 跑完馬上看得到
    print(f"\n{'=' * 52}", file=sys.stderr)
    for item in summary:
        note = f"  ({item['note']})" if item["note"] else ""
        print(f"  {item['company']:<12} {item['count']:>4} 筆{note}", file=sys.stderr)
    print(f"  {'合計':<11} {len(all_jobs):>4} 筆", file=sys.stderr)
    print("=" * 52, file=sys.stderr)

    all_jobs = sort_by_posted(all_jobs)     # 跨公司合併後重新排, 否則會按公司分群

    if not all_jobs:
        print("沒有符合條件的職缺", file=sys.stderr)
        return
    if args.no_save or args.separate:
        return

    name = "workday" if len(urls) > 1 else summary[0]["company"]
    path = args.out or default_output_path(name, taiwan, fmt)
    _write(all_jobs, path, fmt, summary if len(urls) > 1 else None)

    if args.json_out:
        with open(args.json_out, "w", encoding="utf-8") as f:
            json.dump(all_jobs, f, ensure_ascii=False, indent=2)
        print(f"已存檔 JSON -> {os.path.abspath(args.json_out)}", file=sys.stderr)


def _write(jobs, path, fmt, summary):
    """依格式存檔, xlsx 失敗就自動降級成 csv。"""
    if fmt == "xlsx":
        if not save_xlsx(jobs, path, summary):
            print("找不到 openpyxl, 改存 CSV (要 xlsx 請執行: pip install openpyxl)",
                  file=sys.stderr)
            path = os.path.splitext(path)[0] + ".csv"
            save_csv(jobs, path)
    else:
        save_csv(jobs, path)
    print(f"已存檔 {len(jobs)} 筆 -> {os.path.abspath(path)}", file=sys.stderr)


if __name__ == "__main__":
    main()
